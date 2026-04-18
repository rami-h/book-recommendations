#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""ייצוא local_library.js לקובץ XLSX מלא (כל השדות)."""
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\local_library.js'
OUT = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\catalog_updated.xlsx'

with open(SRC, 'r', encoding='utf-8') as f:
    content = f.read()
books = json.loads(content[content.index('['):content.rindex(']')+1])

# סדר עמודות רצוי
COLUMNS = [
    ('id', 'מזהה'),
    ('title', 'כותרת'),
    ('author', 'מחבר'),
    ('genres', 'ז\'אנרים'),
    ('sub_genre', 'תת־ז\'אנר'),
    ('themes', 'נושאים'),
    ('mood', 'מצב רוח'),
    ('style', 'סגנון'),
    ('audience', 'קהל יעד'),
    ('language', 'שפה'),
    ('era', 'תקופה'),
    ('year', 'שנה'),
    ('description', 'תקציר'),
    ('similar_to', 'דומה ל'),
    ('in_library', 'בספרייה'),
    ('library_notes', 'הערות ספרייה'),
    ('shelf', 'מדף'),
    ('series', 'סדרה'),
    ('publisher', 'הוצאה'),
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'קטלוג ספריית שחרות'
ws.sheet_view.rightToLeft = True

# כותרות
header_fill = PatternFill(start_color='1F3D2E', end_color='1F3D2E', fill_type='solid')
header_font = Font(bold=True, color='FBF7ED', size=12, name='Arial')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
right = Alignment(horizontal='right', vertical='top', wrap_text=True)
thin = Side(border_style='thin', color='D9CFBC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col_idx, (key, label) in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=label)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

# שורות
def normalize(value):
    if value is None:
        return ''
    if isinstance(value, list):
        return ', '.join(str(v) for v in value)
    if isinstance(value, bool):
        return 'כן' if value else 'לא'
    return str(value)

for row_idx, book in enumerate(books, 2):
    for col_idx, (key, _) in enumerate(COLUMNS, 1):
        value = normalize(book.get(key, ''))
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = right
        cell.border = border
        cell.font = Font(name='Arial', size=10)

# רוחב עמודות
widths = {
    'id': 10, 'title': 30, 'author': 22, 'genres': 20, 'sub_genre': 18,
    'themes': 25, 'mood': 15, 'style': 15, 'audience': 14, 'language': 12,
    'era': 14, 'year': 8, 'description': 60, 'similar_to': 20,
    'in_library': 12, 'library_notes': 25, 'shelf': 12, 'series': 20, 'publisher': 20,
}
for col_idx, (key, _) in enumerate(COLUMNS, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = widths.get(key, 15)

# גובה שורות דינמי לתקצירים
for row_idx in range(2, len(books) + 2):
    ws.row_dimensions[row_idx].height = 45

# הקפאת כותרת
ws.freeze_panes = 'A2'

wb.save(OUT)

total = len(books)
with_desc = sum(1 for b in books if b.get('description'))
print(f'סה"כ ספרים: {total}')
print(f'עם תקציר:  {with_desc} ({100*with_desc/total:.1f}%)')
print(f'נשמר:     {OUT}')
