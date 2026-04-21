#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
scrape_new_books.py
===================
מגרד תקצירים מסימניה ו-e-vrit.
תומך בריצה על ספרים חדשים (NEW_IDS) או על כל ספרי המבוגרים ללא תקציר.

שימוש:
    py -3 scrape_new_books.py --test              # 5 ספרים, הצג ציון + URL
    py -3 scrape_new_books.py                     # ספרים מ-NEW_IDS
    py -3 scrape_new_books.py --auto-ids          # IDs גבוהים ללא תקציר
    py -3 scrape_new_books.py --all-adults        # כל 906 ספרי המבוגרים ללא תקציר
    py -3 scrape_new_books.py --all-adults --resume  # המשך מ-checkpoint
"""

import json, re, sys, io, time, os, argparse, urllib.parse
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ADULTS_PATH = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\local_library_adults.js'
KIDS_PATH   = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\local_library_kids.js'
CKPT_PATH   = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\scrape_adults_checkpoint.jsonl'
BORDER_XLSX = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\borderline_review.xlsx'

# IDs של הספרים החדשים מה-PDF — מעדכנים ידנית אחרי כל קליטה
NEW_IDS = {f'lib{i}' for i in list(range(5781, 5826)) + [3133]}

SIMANIA_DELAY  = 2.5
SAVE_EVERY     = 25

# ספי קבלה — recall score
THRESHOLD_ACCEPT = 0.65   # קבל אוטומטית
THRESHOLD_BORDER = 0.45   # שמור לבדיקה ידנית

# נרמול שמות מחברים — וריאציות ידועות → צורת חיפוש מיטבית
AUTHOR_NORMALIZE = {
    'זיס דוקטור':  'דוקטור סוס',
    'ד"ר סאוס':    'דוקטור סוס',
    'ד"ר סוס':     'דוקטור סוס',
    'דוקטור זוס':  'דוקטור סוס',
    'סאוס ד"ר':    'דוקטור סוס',
    'סוס ד"ר':     'דוקטור סוס',
}


# ── I/O ──────────────────────────────────────────────────────────────────────

def load_js(path):
    with open(path, encoding='utf-8-sig') as f:
        content = f.read()
    m = re.search(r'var LOCAL_LIBRARY\w*\s*=\s*(\[.*\])\s*;', content, re.DOTALL)
    return json.loads(m.group(1)), content, m


def save_js(path, books, content, m):
    with open(path, 'w', encoding='utf-8') as f:
        f.write(content[:m.start(1)])
        f.write(json.dumps(books, ensure_ascii=False, indent=2))
        f.write(content[m.end(1):])


def load_checkpoint():
    done = {}
    if not os.path.exists(CKPT_PATH):
        return done
    with open(CKPT_PATH, encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    e = json.loads(line)
                    done[e['id']] = e
                except Exception:
                    pass
    return done


def append_checkpoint(entry):
    with open(CKPT_PATH, 'a', encoding='utf-8') as f:
        f.write(json.dumps(entry, ensure_ascii=False) + '\n')


# ── כלי עזר לדמיון ────────────────────────────────────────────────────────────

def normalize_word(w):
    """מסיר ה' הידיעה מתחילת מילה לצורך השוואה."""
    return w[1:] if len(w) > 2 and w[0] == 'ה' else w


def search_title(title):
    """כותרת ראשית בלבד — מסיר תת-כותרת אחרי נקודותיים או מקף."""
    return re.split(r'\s*[:\u2014\u2013]\s*', title)[0].strip()


def recall_score(our_title, their_title):
    """כמה מהמילים שלנו מופיעות בכותרת שלהם (recall), מינוס עונש קטן על מילים זרות."""
    our   = {normalize_word(w) for w in re.sub(r'[^\w\s]', '', search_title(our_title)).split()}
    their = {normalize_word(w) for w in re.sub(r'[^\w\s]', '', their_title).split()}
    if not our:
        return 0.0
    rc = len(our & their) / len(our)
    extra = their - our
    return rc - min(0.15, len(extra) * 0.02)


def normalize_author(author):
    return AUTHOR_NORMALIZE.get(author.strip(), author) if author else author


def author_search_key(author):
    """שם משפחה (מילה אחרונה) כמפתח חיפוש."""
    norm = normalize_author(author)
    parts = norm.split() if norm else []
    return parts[-1] if parts else ''


# ── fetchers ──────────────────────────────────────────────────────────────────

def _simania_search(page, query):
    """מבצע חיפוש בסימניה ומחזיר רשימת קישורים."""
    search_url = 'https://simania.co.il/searchBooks.php?query=' + urllib.parse.quote(query)
    try:
        page.goto(search_url, wait_until='domcontentloaded', timeout=15000)
        time.sleep(1.5)
    except Exception as e:
        print(f'  [SIM NAV ERR] {str(e)[:60]}')
        return []
    return page.evaluate('''() => {
        const links = Array.from(document.querySelectorAll('a[href*="bookdetails.php"]'));
        const seen = new Set();
        const result = [];
        for (const a of links) {
            const m = a.href.match(/item_id=(\\d+)/);
            if (!m) continue;
            const id = m[1];
            if (seen.has(id)) continue;
            const text = (a.textContent || '').trim();
            if (text.length < 2 || text.includes('מוכרים')) continue;
            seen.add(id);
            result.push({ id, href: 'https://simania.co.il/bookdetails.php?item_id=' + id, text });
            if (result.length >= 10) break;
        }
        return result;
    }''')


def fetch_simania(page, title, author):
    """
    מחזיר (desc, matched_title, url, score).
    desc מוגדר רק אם score >= THRESHOLD_ACCEPT.
    url מוגדר גם עבור מקרים גבוליים.
    """
    author_key  = author_search_key(author)
    norm_author = normalize_author(author) if author else ''
    st          = search_title(title)

    def best_score_for(links):
        if not links:
            return 0.0
        return max(recall_score(title, l['text']) for l in links)

    # ניסיון 1: כותרת ראשית + שם משפחה מחבר
    query1 = f'{st} {author_key}'.strip()
    book_links = _simania_search(page, query1)

    # ניסיון 2 (אם ציון נמוך מדי): שתי המילים הראשונות של הכותרת
    if best_score_for(book_links) < THRESHOLD_BORDER:
        first_two = ' '.join(st.split()[:2])
        if first_two and first_two != query1:
            cand = _simania_search(page, first_two)
            if best_score_for(cand) > best_score_for(book_links):
                book_links = cand

    # ניסיון 3 (אם עדיין ציון נמוך): שם מחבר מלא (לספרים עם כותרת בעייתית כמו "המלך צב-צב")
    if best_score_for(book_links) < THRESHOLD_BORDER and norm_author and norm_author != author_key:
        cand = _simania_search(page, norm_author)
        if best_score_for(cand) > best_score_for(book_links):
            book_links = cand

    if not book_links:
        return None, None, None, 0.0

    best  = max(book_links, key=lambda l: recall_score(title, l['text']))
    score = recall_score(title, best['text'])

    if score < THRESHOLD_BORDER:
        return None, None, None, 0.0

    # גבולי או קביל — ממשיך לדף הספר
    try:
        page.goto(best['href'], wait_until='domcontentloaded', timeout=15000)
        time.sleep(1.5)
    except Exception as e:
        print(f'  [SIM BOOK ERR] {str(e)[:60]}')
        return None, best['text'], best['href'], score

    desc = page.evaluate('''() => {
        const all = Array.from(document.querySelectorAll('div, p'));
        const blocks = [];
        const seen = new Set();
        for (const el of all) {
            if (el.closest('.review-preview-text') || el.classList.contains('review-preview-text')) continue;
            if (el.querySelectorAll('div, p').length > 2) continue;
            const text = (el.textContent || '').trim();
            if (text.length < 120 || text.length > 4000) continue;
            if (text.startsWith('הספר ') && text.includes('יצא לאור')) continue;
            const sig = text.substring(0, 80);
            if (seen.has(sig)) continue;
            seen.add(sig);
            blocks.push(text);
            if (blocks.length >= 3) break;
        }
        return blocks[0] || null;
    }''')

    if not desc or len(desc) < 80:
        return None, best['text'], best['href'], score

    desc = re.sub(r'\s+', ' ', desc).strip()
    desc = re.sub(r'\s*(קרא עוד|קראו עוד|להמשך קריאה|עוד\.\.\.?)\s*$', '', desc).strip()
    if len(desc) > 2500:
        cut = desc[:2500]
        m2  = re.search(r'[.!?](?=[^.!?]*$)', cut)
        desc = cut[:m2.start()+1].strip() if m2 and m2.start() > 1500 else cut + '...'

    if score >= THRESHOLD_ACCEPT:
        return desc, best['text'], best['href'], score
    else:
        # גבולי — יש תקציר אבל ציון נמוך, שמור לבדיקה
        return None, best['text'], best['href'], score


def fetch_evrit(page, title, author):
    """
    מחזיר (desc, matched_title, url, score).
    desc מוגדר רק אם score >= THRESHOLD_ACCEPT.
    """
    search_url = 'https://www.e-vrit.co.il/Search/' + urllib.parse.quote(search_title(title))
    try:
        page.goto(search_url, wait_until='domcontentloaded', timeout=25000)
        time.sleep(7.0)
    except Exception as e:
        print(f'  [EVR NAV ERR] {str(e)[:60]}')
        return None, None, None, 0.0

    links = page.evaluate('''() => {
        const container = document.querySelector('[class*="product-list"]');
        if (!container) return [];
        const seen = new Set();
        const out = [];
        container.querySelectorAll('a[href*="/Product/"]').forEach(a => {
            const m = a.href.match(/Product\/(\\d+)\//);
            if (!m) return;
            const id = m[1];
            if (seen.has(id)) return;
            seen.add(id);
            const img = a.querySelector('img') || a.closest('[class]')?.querySelector('img');
            const alt = img?.alt?.trim() || '';
            const slug = decodeURIComponent(a.href.split('/Product/')[1]?.split('/').slice(1).join(' ') || '').replace(/_/g,' ').trim();
            const t = alt || slug;
            if (t.length < 2) return;
            out.push({ id, href: a.href, title: t });
            if (out.length >= 20) return;
        });
        return out;
    }''')

    if not links:
        return None, None, None, 0.0

    def evrit_score(link):
        our   = {normalize_word(w) for w in re.sub(r'[^\w\s]', '', search_title(title)).split()}
        their = {normalize_word(w) for w in re.sub(r'[^\w\s]', '', link['title']).split()}
        if not our: return 0.0
        rc    = len(our & their) / len(our)
        extra = their - our
        return rc - min(0.15, len(extra) * 0.02)

    best  = max(links, key=evrit_score)
    score = evrit_score(best)

    if score < THRESHOLD_BORDER:
        return None, None, None, 0.0

    # ממשיך לדף הספר (גם עבור גבוליים — כדי לשמור תקציר ב-Excel)
    try:
        page.goto(best['href'], wait_until='domcontentloaded', timeout=20000)
        time.sleep(3.0)
    except Exception as e:
        print(f'  [EVR BOOK ERR] {str(e)[:60]}')
        return None, best['title'], best['href'], score

    data = page.evaluate('''() => {
        // ניסיון 1: JSON-LD schema (אמין ביותר)
        const scripts = Array.from(document.querySelectorAll('script[type="application/ld+json"]'));
        for (const s of scripts) {
            try {
                const obj = JSON.parse(s.textContent);
                const arr = Array.isArray(obj) ? obj : [obj];
                for (const item of arr) {
                    if (item.description && item.description.length > 60) {
                        return { desc: item.description };
                    }
                }
            } catch(e) {}
        }
        // ניסיון 2: פסקאות p:not([class])
        const skipWords = ['קנייה','לסל','מחיר','שקל','נרשם','התחבר','עוגי',
                           'בתמונה:','הוצאה:','תאריך הוצאה:','מספר עמודים:','קטגוריה:'];
        const paras = Array.from(document.querySelectorAll('p:not([class])'));
        const seen = new Set();
        const parts = [];
        for (const p of paras) {
            const t = (p.textContent || '').trim();
            if (t.length < 40 || t.length > 2000) continue;
            if (seen.has(t.substring(0, 40))) continue;
            if (skipWords.some(w => t.includes(w))) continue;
            seen.add(t.substring(0, 40));
            parts.push(t);
            if (parts.join(' ').length > 1800) break;
        }
        return { desc: parts.join('\\n') };
    }''')

    desc = (data.get('desc') or '').strip()
    if not desc or len(desc) < 80:
        return None, best['title'], best['href'], score

    desc = re.sub(r'\s+', ' ', desc).strip()

    if score >= THRESHOLD_ACCEPT:
        return desc, best['title'], best['href'], score
    else:
        return None, best['title'], best['href'], score


# ── זיהוי אוטומטי ─────────────────────────────────────────────────────────────

def auto_detect_new_ids(adults, kids, top_n=60):
    all_books = adults + kids
    def id_num(b):
        m = re.search(r'\d+', b['id'])
        return int(m.group()) if m else 0
    sorted_books = sorted(all_books, key=id_num, reverse=True)
    no_desc = [b for b in sorted_books if not b.get('description', '').strip()]
    return {b['id'] for b in no_desc[:top_n]}


# ── ייצוא גבוליים ─────────────────────────────────────────────────────────────

def export_borderline(cases):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('openpyxl לא מותקן — לא ייצא Excel')
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'מקרים גבוליים'
    ws.sheet_view.rightToLeft = True

    headers = ['ID', 'כותרת שלנו', 'כותרת שנמצאה', 'ציון', 'מקור', 'לינק', 'תקציר (קטע)', 'אישור (כן/לא)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='2F5496')
        cell.alignment = Alignment(horizontal='center')

    for row, c in enumerate(cases, 2):
        ws.cell(row=row, column=1, value=c['id'])
        ws.cell(row=row, column=2, value=c['our_title'])
        ws.cell(row=row, column=3, value=c.get('matched', ''))
        ws.cell(row=row, column=4, value=round(c.get('score', 0), 2))
        ws.cell(row=row, column=5, value=c.get('source', ''))
        url_cell = ws.cell(row=row, column=6, value=c.get('url', ''))
        url_cell.hyperlink = c.get('url', '')
        url_cell.font = Font(color='0563C1', underline='single')
        snippet = (c.get('desc') or '')[:150]
        ws.cell(row=row, column=7, value=snippet)
        ws.cell(row=row, column=8, value='')
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill('solid', start_color='DCE6F1')

    for col, w in zip('ABCDEFGH', [10, 36, 32, 7, 9, 12, 50, 12]):
        ws.column_dimensions[col].width = w

    wb.save(BORDER_XLSX)
    print(f'\nגבוליים נשמרו: {BORDER_XLSX}  ({len(cases)} מקרים)')


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--test',       action='store_true', help='5 ספרים ראשונים בלבד')
    parser.add_argument('--auto-ids',   action='store_true', help='זיהוי לפי IDs גבוהים')
    parser.add_argument('--all-adults', action='store_true', help='כל ספרי המבוגרים ללא תקציר')
    parser.add_argument('--resume',     action='store_true', help='המשך מ-checkpoint')
    args = parser.parse_args()

    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print('Playwright לא מותקן. הרץ: py -3 -m playwright install chromium')
        sys.exit(1)

    adults, ac, am = load_js(ADULTS_PATH)
    kids,   kc, km = load_js(KIDS_PATH)

    # קבע target_ids
    if args.all_adults:
        target_ids = {b['id'] for b in adults if not b.get('description', '').strip()}
        print(f'מצב --all-adults: {len(target_ids)} ספרי מבוגרים ללא תקציר')
    elif args.auto_ids:
        target_ids = auto_detect_new_ids(adults, kids)
        print(f'מצב --auto-ids: {len(target_ids)} ספרים')
    else:
        target_ids = NEW_IDS

    # checkpoint
    done_ckpt = load_checkpoint() if args.resume else {}
    if args.resume:
        print(f'checkpoint: {len(done_ckpt)} ספרים כבר עובדו')
        # הזרק תקצירים מ-checkpoint
        adults_by_id = {b['id']: b for b in adults}
        kids_by_id   = {b['id']: b for b in kids}
        for bid, entry in done_ckpt.items():
            if entry.get('status') == 'found' and entry.get('desc'):
                target = adults_by_id.get(bid) or kids_by_id.get(bid)
                if target and not target.get('description', '').strip():
                    target['description'] = entry['desc']
                    target['description_source'] = entry.get('source', 'scraped')

    # בנה targets (דלג על checkpoint)
    all_books = [(b, 'adults') for b in adults] + [(b, 'kids') for b in kids]
    targets = [(b, src) for b, src in all_books
               if b['id'] in target_ids
               and not b.get('description', '').strip()
               and b['id'] not in done_ckpt]

    if args.test:
        targets = targets[:5]

    print(f'ספרים לגריפה: {len(targets)}')
    if args.all_adults:
        eta_min = len(targets) * (SIMANIA_DELAY + 8) / 60
        print(f'זמן משוער: ~{eta_min:.0f} דקות')
    print()

    adults_by_id = {b['id']: b for b in adults}
    kids_by_id   = {b['id']: b for b in kids}

    found          = 0
    borderline_cases = []
    not_found      = []

    with sync_playwright() as p:
        browser  = p.chromium.launch(headless=True)
        page_sim = browser.new_page()
        page_evr = browser.new_page()

        for i, (book, src) in enumerate(targets, 1):
            title  = book['title']
            author = book.get('author', '')
            bid    = book['id']
            print(f'[{i}/{len(targets)}] {title[:45]}')

            # שלב 1 — סימניה
            sim_desc, sim_matched, sim_url, sim_score = fetch_simania(page_sim, title, author)

            # שלב 2 — evrit (תמיד ננסה אם סימניה לא קיבלה)
            evr_desc, evr_matched, evr_url, evr_score = None, None, None, 0.0
            if not sim_desc:
                time.sleep(0.5)
                evr_desc, evr_matched, evr_url, evr_score = fetch_evrit(page_evr, title, author)

            # החלטה
            desc    = sim_desc or evr_desc
            matched = sim_matched if sim_desc else evr_matched
            url     = sim_url if sim_desc else evr_url
            source  = 'simania' if sim_desc else 'evrit'
            score   = sim_score if sim_desc else evr_score

            if desc:
                # קביל
                target = adults_by_id.get(bid) or kids_by_id.get(bid)
                if target:
                    target['description'] = desc
                    target['description_source'] = source
                found += 1
                append_checkpoint({'id': bid, 'status': 'found', 'desc': desc,
                                   'matched': matched, 'url': url, 'source': source, 'score': score})
                if args.test:
                    print(f'  ✓ [{source}] {title[:30]} → {(matched or "")[:30]}')
                    print(f'  ציון: {score:.2f} | {url}')
                    print(f'  {desc[:180]}{"..." if len(desc)>180 else ""}')
                else:
                    print(f'  ✓ [{source}] {(matched or "")[:40]} (ציון: {score:.2f})')

            else:
                # בדוק גבולי
                best_score   = max(sim_score, evr_score)
                best_matched = sim_matched if sim_score >= evr_score else evr_matched
                best_url     = sim_url     if sim_score >= evr_score else evr_url
                best_source  = 'simania'  if sim_score >= evr_score else 'evrit'
                # תקציר גבולי מה-fetcher (אם נשמר)
                border_desc  = None  # fetchers כבר שמרו None עבור גבולי

                if best_score >= THRESHOLD_BORDER and best_matched:
                    entry = {'id': bid, 'our_title': title, 'matched': best_matched,
                             'url': best_url, 'source': best_source,
                             'score': best_score, 'desc': border_desc}
                    borderline_cases.append(entry)
                    append_checkpoint({'id': bid, 'status': 'borderline', **entry})
                    print(f'  ? גבולי [{best_source}] {(best_matched or "")[:35]} (ציון: {best_score:.2f})')
                    if args.test and best_url:
                        print(f'  {best_url}')
                else:
                    not_found.append(title)
                    append_checkpoint({'id': bid, 'status': 'not_found'})
                    print(f'  ✗ לא נמצא')

            # שמור כל SAVE_EVERY ספרים (לא בטסט)
            if not args.test and i % SAVE_EVERY == 0:
                save_js(ADULTS_PATH, adults, ac, am)
                save_js(KIDS_PATH,   kids,   kc, km)
                print(f'  [שמירה ביניים — {i} ספרים]')

            time.sleep(SIMANIA_DELAY)

        page_sim.close()
        page_evr.close()
        browser.close()

    # שמירה סופית
    if not args.test:
        save_js(ADULTS_PATH, adults, ac, am)
        save_js(KIDS_PATH,   kids,   kc, km)
        print('\nנשמר.')
    else:
        print('\n--- טסט בלבד, לא נשמר ---')

    # סיכום
    total = len(targets)
    print(f'\nנמצאו:   {found}/{total}')
    print(f'גבוליים: {len(borderline_cases)}/{total}')
    print(f'לא נמצאו: {len(not_found)}/{total}')

    if not_found and (args.test or len(not_found) <= 20):
        print('\nלא נמצאו:')
        for t in not_found:
            print(f'  - {t}')

    # ייצוא גבוליים
    if borderline_cases and not args.test:
        export_borderline(borderline_cases)
    elif borderline_cases and args.test:
        print(f'\nגבוליים (טסט — לא ייצוא):')
        for c in borderline_cases:
            print(f'  ? {c["our_title"][:35]} → {c["matched"][:35]} ({c["score"]:.2f}) {c["url"]}')


if __name__ == '__main__':
    main()
