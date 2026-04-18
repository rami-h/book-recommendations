#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Run simania scraper on books in an xlsx file and write descriptions back."""
import openpyxl
import time
import json
import os
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from scrape_descriptions import fetch_simania
from playwright.sync_api import sync_playwright

SRC = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\catalog_no_match_only.xlsx'
OUT = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\catalog_no_match_scraped.xlsx'
PROGRESS = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\scrape_xlsx_progress.json'
LOG = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\scrape_xlsx_log.txt'

COL_ID = 1
COL_TITLE = 2
COL_AUTHOR = 3
COL_GENRE = 4
COL_DESC = 13
SAVE_EVERY = 25
DELAY = 2.0


def extract_category(page):
    """חילוץ קטגוריה מדף ספר טעון בסימניה."""
    try:
        cats = page.evaluate('''() => {
            const links = Array.from(document.querySelectorAll('a[href*="category.php?categoryId"]'));
            const seen = new Set();
            const out = [];
            for (const a of links) {
                const t = (a.textContent || '').trim();
                if (!t || t.length > 40 || seen.has(t)) continue;
                seen.add(t);
                out.push(t);
            }
            return out;
        }''')
        if cats:
            return ', '.join(cats[:3])
    except Exception:
        pass
    return None


def log(msg):
    with open(LOG, 'a', encoding='utf-8') as f:
        f.write(msg + '\n')
    try:
        print(msg.encode('ascii', 'replace').decode('ascii'), flush=True)
    except Exception:
        pass


def load_progress():
    if os.path.exists(PROGRESS):
        with open(PROGRESS, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'done_ids': [], 'success': 0, 'tried': 0}


def save_progress(p):
    with open(PROGRESS, 'w', encoding='utf-8') as f:
        json.dump(p, f, ensure_ascii=False, indent=2)


def main():
    # Load source xlsx (or resume from out if it exists)
    source = OUT if os.path.exists(OUT) else SRC
    wb = openpyxl.load_workbook(source)
    ws = wb.active
    total = ws.max_row

    prog = load_progress()
    done = set(prog['done_ids'])

    log(f'=== Starting simania scrape: {total} books ===')
    log(f'Already done: {len(done)}, success so far: {prog["success"]}')

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent='Mozilla/5.0 (ShaharutLibBot)')
        page = ctx.new_page()
        page.set_default_timeout(20000)

        for row in range(1, total + 1):
            book_id = ws.cell(row, COL_ID).value
            if not book_id:
                continue
            title = ws.cell(row, COL_TITLE).value or ''
            author = ws.cell(row, COL_AUTHOR).value or ''
            existing_desc = ws.cell(row, COL_DESC).value
            existing_genre = ws.cell(row, COL_GENRE).value
            has_desc = existing_desc and len(str(existing_desc)) > 80
            has_genre = existing_genre and str(existing_genre).strip()

            # דלג אם הכל כבר קיים או אם כבר ניסינו
            if has_desc and has_genre:
                done.add(book_id)
                continue
            if book_id in done:
                continue

            prog['tried'] += 1
            log(f'[{row}/{total}] {title} / {author}')

            desc = None
            category = None
            try:
                result = fetch_simania(page, title, author)
                desc = result[0] if isinstance(result, tuple) else result
                if desc:
                    # page עדיין טעון על דף הספר - נחלץ קטגוריה
                    category = extract_category(page)
            except Exception as e:
                log(f'  ERR: {str(e)[:80]}')

            wrote = []
            if desc and not has_desc:
                ws.cell(row, COL_DESC, value=desc)
                wrote.append(f'desc({len(desc)})')
            if category and not has_genre:
                ws.cell(row, COL_GENRE, value=category)
                wrote.append(f'genre={category}')
            if wrote:
                prog['success'] += 1
                log(f'  OK  {" | ".join(wrote)}')
            else:
                log('  --')

            done.add(book_id)
            prog['done_ids'] = list(done)

            if prog['tried'] % SAVE_EVERY == 0:
                wb.save(OUT)
                save_progress(prog)
                log(f'  [SAVED] success={prog["success"]}/{prog["tried"]}')

            time.sleep(DELAY)

        browser.close()

    wb.save(OUT)
    save_progress(prog)
    log(f'\n=== DONE ===')
    log(f'Tried: {prog["tried"]}')
    log(f'Success: {prog["success"]} ({100*prog["success"]/max(prog["tried"],1):.1f}%)')
    log(f'Output: {OUT}')


if __name__ == '__main__':
    main()
