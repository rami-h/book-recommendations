#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Enrich genre only for rows that already have description but missing genre."""
import openpyxl
import time
import sys

from scrape_descriptions import fetch_simania
from playwright.sync_api import sync_playwright
from scrape_xlsx import extract_category

SRC = r'C:\Users\hacmo\Desktop\MyWebsite\library\data\catalog_no_match_scraped.xlsx'

wb = openpyxl.load_workbook(SRC)
ws = wb.active

# Find rows: has desc, missing genre
targets = []
for r in range(1, ws.max_row + 1):
    d = ws.cell(r, 13).value
    g = ws.cell(r, 4).value
    if d and len(str(d)) > 80 and (not g or not str(g).strip()):
        targets.append((r, ws.cell(r, 2).value, ws.cell(r, 3).value))

print(f'Targets: {len(targets)}')

filled = 0
with sync_playwright() as p:
    browser = p.chromium.launch(headless=True)
    page = browser.new_context(user_agent='Mozilla/5.0').new_page()
    page.set_default_timeout(20000)

    for i, (row, title, author) in enumerate(targets, 1):
        msg = f'[{i}/{len(targets)}] {title} / {author}'
        try:
            print(msg.encode('ascii', 'replace').decode('ascii'), flush=True)
        except Exception:
            pass
        try:
            result = fetch_simania(page, title or '', author or '')
            if result and result[0]:
                cat = extract_category(page)
                if cat:
                    ws.cell(row, 4, value=cat)
                    filled += 1
                    print(f'  OK genre={cat}'.encode('ascii', 'replace').decode('ascii'), flush=True)
                else:
                    print('  -- no category', flush=True)
            else:
                print('  -- no match', flush=True)
        except Exception as e:
            print(f'  ERR: {str(e)[:60]}', flush=True)
        time.sleep(1.5)
    browser.close()

wb.save(SRC)
print(f'\nFilled {filled}/{len(targets)} genres')
